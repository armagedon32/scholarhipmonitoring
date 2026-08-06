"""Convert a real historical-record file (CSV or Excel) into the training CSV.

The school's records may use a simple spreadsheet format such as:
  Name, GWA, Units, Failed Subject, Annual Family Income, Attendance,
  Socio Economic Status, Prediction   (Prediction = Retained / At-Risk)

This converter maps those columns onto the retention model's training schema and
writes `data/scholar_data.csv` so the model can be retrained with real history.

Fields the sheet may not provide are filled with sensible defaults:
  - semester_performance  -> set to GWA (their nearest gauge of term standing)
  - scholarship_type      -> from an optional 'Scholarship'/'Grant' column, else "Academic"
  - academic_year         -> "Historical" (informational; ignored by the model)

Usage:  python -m ml.convert_history --input "path/to/records.xlsx"
        python -m ml.track_model   # after converting, retrains on the real data
"""
import argparse
import csv
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from config import Config
from ml.prep import CATEGORY_LEVELS

SOCIO_SYNONYMS = {
    "low": "Low", "lower": "Lower-Middle", "lower-middle": "Lower-Middle",
    "lower middle": "Lower-Middle", "middle": "Middle", "midd": "Middle",
    "upper-middle": "Upper-Middle", "upper middle": "Upper-Middle", "upper": "Upper-Middle",
}

SCHOLAR_SYNONYMS = {
    "academic": "Academic", "dost": "DOST", "ched": "CHED_GIA",
    "ched-gia": "CHED_GIA", "ched gia": "CHED_GIA", "municipal": "Municipal",
    "muni": "Municipal", "lgu": "Municipal",
}

PREDICTION_MAP = {
    "retained": 1, "retain": 1, "keep": 1, "1": 1,
    "0": 0, "at-risk": 0, "at risk": 0, "risk": 0, "drop": 0,
    "dropped": 0, "ineligible": 0,
}


def _norm(s):
    return (s or "").strip().lower()


def _lookup(header, hints):
    for h in header:
        nh = _norm(h)
        for t in hints:
            if nh == t or nh.startswith(t):
                return h
    return None


def read_rows(path):
    ext = os.path.splitext(path)[1].lower()
    if ext in (".csv", ".txt", ".tsv"):
        sep = "\t" if ext == ".tsv" else ","
        with open(path, newline="", encoding="utf-8-sig") as fh:
            reader = csv.reader(fh, delimiter=sep)
            raw = list(reader)
        return raw[0], raw[1:]
    if ext in (".xlsx", ".xlsm"):
        from openpyxl import load_workbook
        wb = load_workbook(path, data_only=True)
        ws = wb.active
        raw = list(ws.iter_rows(values_only=True))
        header = [str(c) if c is not None else "" for c in raw[0]]
        rows = [list(r) for r in raw[1:]]
        return header, rows
    raise SystemExit("Unsupported file type. Use .csv, .tsv, or .xlsx")


def convert(path, out_path):
    header, rows = read_rows(path)

    c_name = _lookup(header, ["name", "student"])
    c_gwa = _lookup(header, ["gwa", "gpa", "grade average"])
    c_units = _lookup(header, ["units", "enrolled"])
    c_failed = _lookup(header, ["failed subject", "failing", "failed", "fail"])
    c_income = _lookup(header, ["annual family income", "annual income", "family income", "income"])
    c_att = _lookup(header, ["attendance", "attend"])
    c_socio = _lookup(header, ["socio", "socioeconomic", "economic status"])
    c_pred = _lookup(header, ["prediction", "predict", "result", "label", "retention"])
    c_scholar = _lookup(header, ["scholarship", "grant", "scholarship type"])

    missing = [label for label, c in [("GWA", c_gwa), ("Attendance", c_att), ("Prediction", c_pred)] if not c]
    if missing:
        raise SystemExit(f"Could not find required column(s): {', '.join(missing)}.\nHeaders: {header}")

    idx = {}
    for key, h in [("name", c_name), ("gwa", c_gwa), ("units", c_units),
                   ("failed", c_failed), ("income", c_income), ("att", c_att),
                   ("socio", c_socio), ("pred", c_pred), ("scholar", c_scholar)]:
        if h:
            idx[key] = header.index(h)

    def get(r, key):
        i = idx.get(key)
        return str(r[i]).strip() if i is not None and i < len(r) and r[i] is not None else ""

    def num(val, default):
        try:
            return float(val)
        except (TypeError, ValueError):
            return default

    out = []
    skipped = 0
    for r in rows:
        if not any((c is not None and str(c).strip() != "") for c in r):
            continue
        try:
            gwa = float(get(r, "gwa"))
        except ValueError:
            skipped += 1
            continue

        pred = PREDICTION_MAP.get(_norm(get(r, "pred")))
        if pred is None:
            skipped += 1
            continue

        socio = SOCIO_SYNONYMS.get(_norm(get(r, "socio")), "Middle")
        if socio not in CATEGORY_LEVELS["socio_status"]:
            socio = "Middle"
        scholar_type = SCHOLAR_SYNONYMS.get(_norm(get(r, "scholar")), "Academic")
        if scholar_type not in CATEGORY_LEVELS["scholarship_type"]:
            scholar_type = "Academic"

        out.append({
            "student_name": get(r, "name"),
            "academic_year": "Historical",
            "year_level": 0,
            "scholarship_type": scholar_type,
            "socio_status": socio,
            "annual_income": round(num(get(r, "income"), 250000.0), 2),
            "gwa": round(gwa, 2),
            "failed_subjects": int(round(num(get(r, "failed"), 0))),
            "units_enrolled": int(round(num(get(r, "units"), 15))),
            "attendance_rate": round(num(get(r, "att"), 90.0), 1),
            "semester_performance": round(gwa, 2),
            "retained": pred,
        })

    if not out:
        raise SystemExit("No usable rows. Check Prediction values (Retained / At-Risk) and GWA.")

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    with open(out_path, "w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=list(out[0].keys()))
        writer.writeheader()
        writer.writerows(out)

    print(f"Imported {len(out)} records -> {out_path}")
    if skipped:
        print(f"Skipped {skipped} row(s) with missing/invalid GWA or Prediction.")


def main():
    ap = argparse.ArgumentParser(description="Convert school records to training CSV.")
    ap.add_argument("--input", required=True, help="Path to the source CSV/Excel file")
    ap.add_argument("--out", default=None, help="Output training CSV (default data/scholar_data.csv)")
    args = ap.parse_args()
    convert(args.input, args.out or Config.DATASET_PATH)


if __name__ == "__main__":
    main()